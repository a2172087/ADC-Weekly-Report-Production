import sys
sys.stdout.reconfigure(encoding='utf-8')
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import os
import openpyxl
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 新ChromeDriver的路徑
path_to_new_chromedriver = r'D:\Racky\公用相關\暫存檔案\chromedriver-win64\chromedriver.exe'

# 創建ChromeDriver的配置
chrome_options = webdriver.ChromeOptions()
# 在此添加其他選項，例如禁用圖片、使用無頭模式等
chrome_options.add_argument("--disable-images")  # 禁用圖片載入
chrome_options.add_argument("--headless")  # 使用無頭模式

# 創建新的Service物件，使用新的ChromeDriver
new_service = Service(path_to_new_chromedriver)

# 創建新的WebDriver實例，使用新的Service物件和chrome_options
driver = webdriver.Chrome(service=new_service, options=chrome_options)

# 指定文件路徑和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS AVI Output for PAD.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打開工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

# 檢查 X3 單元格內容
cell_value = worksheet['X3'].value
if cell_value != "BUMP OR PAD":
    print("Excel TS AVI Output for PAD V file is no match BUMP or PAD, pls check data!!!")
else:
    # 刪除前三行
    for _ in range(3):
        worksheet.delete_rows(1)

    # 遍歷 X 列，保留包含 "pad" 或爲空（空格）的單元格
    row_index = 1
    while row_index <= worksheet.max_row:
        cell_value = worksheet[get_column_letter(24) + str(row_index)].value
        if cell_value is None or cell_value.strip() == "" or "pad" in cell_value.lower():
            row_index += 1
        else:
            worksheet.delete_rows(row_index)

    # 保存修改後的工作簿
    workbook.save(file_path)

    print("成功建立檔案 TS AVI Output for PAD!")


# 指定文件路徑和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS AVI Output for Bump.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打開工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

# 檢查 X3 單元格內容
cell_value = worksheet['X3'].value
if cell_value != "BUMP OR PAD":
    print("Excel TS AVI Output for Bump V file is no match BUMP or PAD, pls check data!!!")

else:
    # 刪除前三行
    for _ in range(3):
        worksheet.delete_rows(1)

    # 反向遍歷 X 列，保留包含 "bump" 的單元格
    row_index = worksheet.max_row
    while row_index >= 1:
        cell_value = worksheet[get_column_letter(24) + str(row_index)].value
        if cell_value is None or "bump" in cell_value.lower():
            row_index -= 1
        else:
            worksheet.delete_rows(row_index)

    # 保存修改後的工作簿
    workbook.save(file_path)

    # 保存修改後的工作簿
    workbook.save(file_path)

    print("成功建立檔案 TS AVI Output for Bump!")

# 指定文件路徑和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS AVI ADC Output for Pad.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打開工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

# 檢查 X3 單元格內容
cell_value = worksheet['X3'].value
if cell_value != "BUMP OR PAD":
    print("TS AVI ADC Output for Pad X file is no match BUMP or PAD, pls check data!!!")
else:
    # 刪除前三行
    for _ in range(3):
        worksheet.delete_rows(1)

    # 遍歷X列，刪除包含 "bump" 的單元格（不區分大小寫）
    column_v = worksheet['X']
    for cell in column_v:
        if cell.value is None or "bump" in str(cell.value).lower():
            worksheet.delete_rows(cell.row)

    # 保存修改後的工作簿
    workbook.save(file_path)

    print("成功建立檔案 TS AVI ADC Output for Pad!")

# 指定文件路徑和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS AVI ADC Output for Bump.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打開工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active

# 檢查 X3 單元格內容
cell_value = worksheet['X3'].value
if cell_value != "BUMP OR PAD":
    print("TS AVI ADC Output for Bump X file is no match BUMP or PAD, pls check data!!!")
else:
    # 刪除前三行
    for _ in range(3):
        worksheet.delete_rows(1)

    # 遍歷X列，刪除不包含 "bump" 的單元格（不區分大小寫）
    column_v = worksheet['X']
    for cell in column_v:
        if cell.value is None or "bump" not in str(cell.value).lower():
            worksheet.delete_rows(cell.row)

    # 保存修改後的工作簿
    workbook.save(file_path)

    print("成功建立檔案 TS AVI ADC Output for Bump!")

