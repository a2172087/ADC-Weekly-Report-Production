import sys
sys.stdout.reconfigure(encoding='utf-8')
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import glob
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import isoweek
from openpyxl.utils import column_index_from_string
import warnings
from selenium.webdriver.support.ui import Select
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 新ChromeDriver的路徑
path_to_new_chromedriver = r'D:\Racky\公用相關\暫存檔案\chromedriver-win64\chromedriver.exe'

# 創建ChromeDriver的配置
chrome_options = webdriver.ChromeOptions()
# 在此添加其他選項，例如禁用圖片、使用無頭模式等
chrome_options.add_argument("--disable-images")  # 禁用圖片載入
chrome_options.add_argument("--disable-images")  # 禁用圖片載入
chrome_options.add_argument("--headless")  # 使用無頭模式

# 創建新的Service物件，使用新的ChromeDriver
new_service = Service(path_to_new_chromedriver)

# 創建新的WebDriver實例，使用新的Service物件和chrome_options
driver = webdriver.Chrome(service=new_service, options=chrome_options)

url = "http://tstpas/TPAS/index.jsp"
user_id = "A005772"
user_pw = "A11111111111"

# 創建 WebDriver 實例 (以 Chrome 爲例)
s = Service(path_to_new_chromedriver)
driver = webdriver.Chrome(service=s)

# 打開網站
driver.get(url)

# 等待 mainFrame 加載完成
wait = WebDriverWait(driver, 10)
wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "mainFrame")))

# 定位 User ID 和 User PW 輸入框 (根據網頁元素的name等定位)
user_id_input = driver.find_element(By.NAME, "userid")
user_pw_input = driver.find_element(By.NAME, "password")

# 輸入 User ID 和 User PW
user_id_input.send_keys(user_id)
user_pw_input.send_keys(user_pw)

# 提交登錄表單
login_form = driver.find_element(By.NAME, "form1")
login_form.submit()

# 如果需要, 切換回主窗口
driver.switch_to.default_content()

# 切換到包含 'TPAS MFG Utility' 的 frame
frame = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//frame[@src="/TPAS/edu/edu_menu.jsp"]')))
driver.switch_to.frame(frame)

# 點擊 'MFG Utility' 鏈接
xpath = '//a[@id="ygtvlabelel115"]'
element_to_click = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
element_to_click.click()

# 點擊 'MFG Use' 鏈接
xpath = '//a[@id="ygtvlabelel195"]'
element_to_click = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
element_to_click.click()

# 點擊 'Process Output' 鏈接
xpath = '//a[@id="ygtvlabelel210"]'
element_to_click = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
element_to_click.click()

# 將焦點切換回父框架
driver.switch_to.parent_frame()

# 切換到包含 'AVI Inspection Report' 的 frame
frame = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//frame[@src="/TPAS/edu/edu_main.jsp"]')))
driver.switch_to.frame(frame)

# 定位並選擇 "V/M" 和 "IQC" 選項
select_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "pd_name")))
select = Select(select_element)
select.select_by_visible_text("V/M")
select.select_by_visible_text("IQC")

# 計算當周的週四
today = datetime.date.today()
days_to_thursday = (3 - today.weekday()) % 7
this_week_thursday = today + datetime.timedelta(days=days_to_thursday)
this_week_thursday_str = this_week_thursday.strftime("%Y-%m-%d")

# 計算一週前的週四的日期
one_week_ago_thursday = this_week_thursday - datetime.timedelta(days=7)
one_week_ago_thursday_str = one_week_ago_thursday.strftime("%Y-%m-%d")

# 找到名爲 "begin_date" 的輸入框並設置新的日期值
begin_date_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, "begin_date")))
begin_date_input.clear()
begin_date_input.send_keys(one_week_ago_thursday_str)

# 找到名爲 "end_date" 的輸入框並設置新的日期值
end_date_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, "end_date")))
end_date_input.clear()
end_date_input.send_keys(this_week_thursday_str)

# 找到名爲 "Image29" 的按鈕並點擊
submit_button = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, "Image29")))
submit_button.click()

# 暫停 70 秒以觀察網頁操作
time.sleep(70)

# 找到excel下載視窗並點擊
submit_button = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//img[@name='Image13'][@src='/ArdentecW/image/icon_excel.gif']")))
submit_button.click()

# 暫停 15 秒以觀察網頁操作
time.sleep(15)

# 指定下載文件夾
download_folder = r"C:\Users\A005772\Downloads"

# 獲取下載文件夾中的最新文件
list_of_files = glob.glob(download_folder + r'\*')
latest_file = max(list_of_files, key=os.path.getctime)

# 構建新文件名
new_file_name = "TS VM and IQC Output data.xlsx"

# 將下載的文件重命名
shutil.move(latest_file, os.path.join(download_folder, new_file_name))

# 最後，關閉瀏覽器
driver.quit()

# 指定文件路徑和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS VM and IQC Output data.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打開工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sheet = workbook.active

# 遍歷整個工作表的所有單元格
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        
        # 檢查單元格是否是合併儲存格
        is_merged = any(cell.coordinate in merged_cell for merged_cell in sheet.merged_cells)
        
        # 如果單元格是合併儲存格，解除合併
        if is_merged:
            for merged_cell_range in sheet.merged_cells:
                if cell.coordinate in merged_cell_range:
                    sheet.unmerge_cells(str(merged_cell_range))
                    break

# 獲取 A 列的列索引
column_index = column_index_from_string('A')
rows_to_delete = []

# 遍歷 A 列的所有單元格
for row in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=column_index).value
    # 如果值可以轉換為整數，進行轉換
    try:
        cell_value = int(cell_value)
    except (ValueError, TypeError):
        pass
    # 檢查單元格值是否為整數，並在範圍 1 到 9999 之內
    if not (isinstance(cell_value, int) and 1 <= cell_value <= 9999):
        rows_to_delete.append(row)

# 逆序刪除行，避免因行索引改變而導致的錯誤
for row in sorted(rows_to_delete, reverse=True):
    sheet.delete_rows(row)

# 保存修改後的工作簿
workbook.save(file_path)

# 指定文件路徑和文件名
folder_path = r"C:\Users\A005772\Downloads"
old_file_name = "TS VM and IQC Output data.xlsx"
new_file_name_pad = "TS VM and IQC Output data for PAD產品.xlsx"
new_file_name_bump = "TS VM and IQC Output data for Bump產品.xlsx"

# 計算文件完整路徑
old_file_path = os.path.join(folder_path, old_file_name)
new_file_path_pad = os.path.join(folder_path, new_file_name_pad)
new_file_path_bump = os.path.join(folder_path, new_file_name_bump)

# 將指定文件夾"TS VM and IQC Output data.xlsx" 更名為"TS VM and IQC Output data for PAD產品.xlsx"
os.rename(old_file_path, new_file_path_pad)

# 將"TS VM and IQC Output data for PAD產品.xlsx"複製一份檔案並將檔名建立為"TS VM and IQC Output data for Bump產品.xlsx"
shutil.copy(new_file_path_pad, new_file_path_bump)

# 指定文件路径和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS VM and IQC Output data for PAD產品.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打开工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sheet = workbook.active

# 遍历AA列，保留包含 "pad" 或为空（空格）的单元格
row_index = 1
while row_index <= worksheet.max_row:
    cell_value = worksheet[get_column_letter(27) + str(row_index)].value
    if cell_value is None or cell_value.strip() == "" or "pad" in cell_value.lower():
        row_index += 1
    else:
        worksheet.delete_rows(row_index)

# 保存修改后的工作簿
workbook.save(file_path)
print("成功建立檔案 TS VM and IQC Output data for PAD產品")

# 指定文件路径和文件名
folder_path = r"C:\Users\A005772\Downloads"
file_name = "TS VM and IQC Output data for Bump產品.xlsx"
file_path = os.path.join(folder_path, file_name)

# 打开工作簿
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sheet = workbook.active

# 遍歷 AA 列，保留包含 "bump" 的單元格
row_index = 1
while row_index <= worksheet.max_row:
    cell_value = worksheet[get_column_letter(27) + str(row_index)].value
    if cell_value is not None and "bump" in cell_value.lower():
        row_index += 1
    else:
        worksheet.delete_rows(row_index)

# 保存修改后的工作簿
workbook.save(file_path)
print("成功建立檔案 TS VM and IQC Output data for Bump產品")







# 指定文件名和路徑
file_name = "TS AVI Output for Pad.xlsx"
download_folder = r"C:\Users\A005772\Downloads"
file_path = os.path.join(download_folder, file_name)

# 讀取 Excel 文件中 A 列包含文字的行數
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
rows_with_text = sum(1 for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1) if row[0].value)

# 檢查是否存在 ADC Utilization-v2 文件
adc_utilization_file = r"M:\ADC_project\00 3 sites ADC monitor report\24W_ADC Utilization-v2.xlsx"
if not os.path.exists(adc_utilization_file):
    print("No found ADC Utilization-v2 file, pls to check!!")
else:
    # 讀取 ADC Utilization-v2 文件中的所有工作表名
    adc_workbook = openpyxl.load_workbook(adc_utilization_file)
    sheet_names = adc_workbook.sheetnames

    # 選擇相應的工作表並在 F4 單元格中輸入 A 列包含文字的行數
    if 'Week' in sheet_names:
        adc_worksheet = adc_workbook['Week']
        adc_worksheet['F4'] = rows_with_text
        adc_workbook.save(adc_utilization_file)
        print(f"成功更新了 Week 工作表的 F4 單元格")
    else:
        print(f"沒有找到 Week 工作表, 請協助確認檔案 Utilization-v2工作表是否包含 Week 工作表")

# 指定文件名和路徑
file_name = "TS AVI Output for Bump.xlsx"
download_folder = r"C:\Users\A005772\Downloads"
file_path = os.path.join(download_folder, file_name)

# 讀取 Excel 文件中 A 列包含文字的行數
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
rows_with_text_1 = sum(1 for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1) if row[0].value)

# 檢查是否存在 ADC Utilization-v2 文件
adc_utilization_file = r"M:\ADC_project\00 3 sites ADC monitor report\24W_ADC Utilization-v2.xlsx"
if not os.path.exists(adc_utilization_file):
    print("No found ADC Utilization-v2 file, pls to check!!")
else:
    # 讀取 ADC Utilization-v2 文件中的所有工作表名
    adc_workbook = openpyxl.load_workbook(adc_utilization_file)
    sheet_names = adc_workbook.sheetnames

    # 選擇相應的工作表並在 F5 單元格中輸入 A 列包含文字的行數
    if 'Week' in sheet_names:
        adc_worksheet = adc_workbook['Week']
        adc_worksheet['F5'] = rows_with_text_1
        adc_workbook.save(adc_utilization_file)
        print(f"成功更新了 Week 工作表的 F5 單元格")
    else:
        print(f"沒有找到 Week 工作表, 請協助確認檔案 Utilization-v2工作表是否包含 Week 工作表")

# 指定下載文件夾路徑
download_folder = r"C:\Users\A005772\Downloads"

# 指定文件名
file_name = "TS AVI ADC Output for Pad.xlsx"
file_path = os.path.join(download_folder, file_name)

# 讀取 Excel 文件中 E 列的所有數字，並將其相加
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sum_of_e = sum(cell.value for cell in worksheet['E'] if isinstance(cell.value, (int, float)))

# 指定 ADC Utilization-v2 文件的路徑
adc_utilization_file = r"M:\ADC_project\00 3 sites ADC monitor report\24W_ADC Utilization-v2.xlsx"

# 讀取 ADC Utilization-v2 文件
adc_workbook = openpyxl.load_workbook(adc_utilization_file)

# 檢查是否存在名為 "Week" 的工作表
if "Week" in adc_workbook.sheetnames:
    adc_worksheet = adc_workbook["Week"]
    adc_worksheet['G4'] = sum_of_e
    adc_workbook.save(adc_utilization_file)
    print("成功更新了 Week 工作表的 G4 單元格")
else:
    print("沒有找到名為 Week 的工作表，請確認文件是否正確")

# 指定下載文件夾路徑
download_folder = r"C:\Users\A005772\Downloads"

# 指定文件名
file_name = "TS AVI ADC Output for Bump.xlsx"
file_path = os.path.join(download_folder, file_name)

# 讀取 Excel 文件中 E 列的所有數字，並將其相加
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sum_of_e = sum(cell.value for cell in worksheet['E'] if isinstance(cell.value, (int, float)))

# 指定 ADC Utilization-v2 文件的路徑
adc_utilization_file = r"M:\ADC_project\00 3 sites ADC monitor report\24W_ADC Utilization-v2.xlsx"

# 讀取 ADC Utilization-v2 文件
adc_workbook = openpyxl.load_workbook(adc_utilization_file)

# 檢查是否存在名為 "Week" 的工作表
if "Week" in adc_workbook.sheetnames:
    adc_worksheet = adc_workbook["Week"]
    adc_worksheet['G5'] = sum_of_e
    adc_workbook.save(adc_utilization_file)
    print("成功更新了 Week 工作表的 G5 單元格")
else:
    print("沒有找到名為 Week 的工作表，請確認文件是否正確")

# 指定下載文件夾路徑
download_folder = r"C:\Users\A005772\Downloads"

# 指定文件名
file_name = "TS VM and IQC Output data for PAD產品.xlsx"
file_path = os.path.join(download_folder, file_name)

# 讀取 Excel 文件中 N 列的所有數字，並將其相加
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sum_of_e = sum(cell.value for cell in worksheet['N'] if isinstance(cell.value, (int, float)))

# 指定 ADC Utilization-v2 文件的路徑
adc_utilization_file = r"M:\ADC_project\00 3 sites ADC monitor report\24W_ADC Utilization-v2.xlsx"

# 讀取 ADC Utilization-v2 文件
adc_workbook = openpyxl.load_workbook(adc_utilization_file)

# 檢查是否存在名為 "Week" 的工作表
if "Week" in adc_workbook.sheetnames:
    adc_worksheet = adc_workbook["Week"]
    adc_worksheet['F12'] = sum_of_e
    adc_workbook.save(adc_utilization_file)
    print("成功更新了 Week 工作表的 F12 單元格")
else:
    print("沒有找到名為 Week 的工作表，請確認文件是否正確")

# 指定下載文件夾路徑
download_folder = r"C:\Users\A005772\Downloads"

# 指定文件名
file_name = "TS VM and IQC Output data for Bump產品.xlsx"
file_path = os.path.join(download_folder, file_name)

# 讀取 Excel 文件中 N 列的所有數字，並將其相加
workbook = openpyxl.load_workbook(file_path)
worksheet = workbook.active
sum_of_e = sum(cell.value for cell in worksheet['N'] if isinstance(cell.value, (int, float)))

# 指定 ADC Utilization-v2 文件的路徑
adc_utilization_file = r"M:\ADC_project\00 3 sites ADC monitor report\24W_ADC Utilization-v2.xlsx"

# 讀取 ADC Utilization-v2 文件
adc_workbook = openpyxl.load_workbook(adc_utilization_file)

# 檢查是否存在名為 "Week" 的工作表
if "Week" in adc_workbook.sheetnames:
    adc_worksheet = adc_workbook["Week"]
    adc_worksheet['F13'] = sum_of_e
    adc_workbook.save(adc_utilization_file)
    print("成功更新了 Week 工作表的 F13 單元格")
else:
    print("沒有找到名為 Week 的工作表，請確認文件是否正確")

 #獲取當前日期所在的周
now = datetime.datetime.now()
current_year = now.year
current_week = now.isocalendar()[1]
week_str = f"{str(current_year)[-2:]}W{current_week:02d}"

# 原始檔案路徑
original_file_path = "M:\\ADC_project\\00 3 sites ADC monitor report\\24W_ADC Utilization-v2.xlsx"

# 讀取Excel檔案
workbook = openpyxl.load_workbook(original_file_path)

# 直接儲存變更至原始檔案
workbook.save(original_file_path)
